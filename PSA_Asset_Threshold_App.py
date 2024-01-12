import tkinter as tk
import pandas as pd
import PSA_SND_Constants as const
import PSA_SND_Utilities as ut
import PSA_analysis as pa

import openpyxl
import os

from datetime import timedelta, datetime
from tkinter import ttk, messagebox
from tkinter import filedialog as fd
from PIL import Image, ImageTk


class ExcelEntry(tk.Frame):
    def __init__(self, parent, type='file'):
        super().__init__(parent)
        self.dict_config = ut.PSA_SND_read_config(const.strConfigFile)
        self.pic_excel_logo = tk.PhotoImage(file=self.dict_config[const.strWorkingFolder] + "/tkinterapp/assets/Excel_Logo.png").subsample(120)

        if type=='file':
            self.browse1_btn = ttk.Button(self, text="Browse", command=lambda: self.fileBtn_clicked_TB(self.dir_box, '', self.editdata1_btn, True))
            self.editdata1_btn = ttk.Button(self, text="Edit Data", command=lambda: ut.open_xl(self.dir_box, self.dict_config[const.strWorkingFolder]), image=self.pic_excel_logo, compound=tk.LEFT, state=tk.DISABLED)
            self.editdata1_btn.grid(column=2, row=1, padx=10)
            width = 40
        elif type=='folder':
            self.browse1_btn = ttk.Button(self, text="Browse", command=lambda: self.folderBtn_clicked(self.dir_box, '', False))
            width = 55
        else:
            raise ValueError('Invalid type')

        self.browse1_btn.grid(column=0, row=1, padx=10, ipadx=20)

        self.dir_box = tk.Text(self, width=width, height=1)
        self.dir_box.insert('1.0', const.strNoFileDisp)
        self.dir_box['state'] = 'disabled'
        self.dir_box.grid(column=1, row=1, padx=10)

    def fileBtn_clicked_TB(self, fLabel, fpat, fBtn, bInput):
        ftypes = [('Excel files', '.xlsx')]
        fLabel['state'] = 'normal'
        fLabel.delete("1.0", "end")
        fLabel.insert("1.0", const.strNoFileDisp)
        fLabel['state'] = 'disabled'
        
        if bInput:
            file_loc = self.dict_config[const.strWorkingFolder] + '/data/input'
        else:
            file_loc = self.dict_config[const.strWorkingFolder] + '/data/results'
        
        filename = fd.askopenfilename(title='Select file', initialdir=file_loc, filetypes=ftypes, initialfile=fpat)
        fnme_extnd = filename.replace(self.dict_config[const.strWorkingFolder], '')

        if filename != '':
            fLabel['state'] = 'normal'
            fLabel.delete("1.0", "end")
            fLabel.insert("1.0", fnme_extnd)
            fLabel['state'] = 'disabled'
        if bool(fBtn):
            if (fLabel.get("1.0", "end").strip() != const.strNoFileDisp):
                fBtn['state'] = 'normal'
            else:
                fBtn['state'] = 'disabled'

    def folderBtn_clicked(self, fLabel, fpat, bInput):
        fLabel['state'] = 'normal'
        fLabel.delete("1.0", "end")
        fLabel.insert("1.0", const.strNoFileDisp)
        #fLabel['state'] = 'disabled'
    
        if bInput:
            file_loc = self.dict_config[const.strWorkingFolder] + const.strDataInput
        else:
            file_loc = self.dict_config[const.strWorkingFolder] + const.strDataResults

        fldr = fd.askdirectory(title="Select PSA folder", initialdir=file_loc)
        fldr_extnd = fldr.replace(self.dict_config[const.strWorkingFolder], '')
        if fldr != '':
            #fLabel['state'] = 'normal'
            fLabel.delete("1.0", "end")
            fLabel.insert("1.0", fldr_extnd)
            fLabel['state'] = 'disabled'

# -------------------------------------------------------------------------------------------------------------------------------------------------------------------------
'''Main App'''
class App(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        '''Data Entry Parameters'''
        self.title('PSA Threshold Generator')
        self.protocol("WM_DELETE_WINDOW", self.quit_me) 
        # Because of an issue with matplotlib, the program would actually keep running after the exit button is closed while a graph is displayed
        # To counter this, we set protocol to both quit and destroy when the exit button is pressed. 
        
        self.dict_config = ut.PSA_SND_read_config(const.strConfigFile)
        asset_value = tk.StringVar()
        self.df = []
        self.threshold_to_be_used = ''

        self.df_backup = []
        self.i = 0 # Used for undo function
        self.file = ''

        
        LabelFrame00 = ttk.LabelFrame(self, text='Select Files') # File Entry
        LabelFrame10 = ttk.LabelFrame(self, text='Input Parameters') # Data Entry
        LabelFrame20 = ttk.LabelFrame(self, text='Resulting Graph') # Plot Show
        # -----------------------------------------------------------
        '''Frame 0: Select PSA File'''
        frame0 = ttk.LabelFrame(LabelFrame00, text='Select PSA Folder')
    
        self.fe1 = ExcelEntry(frame0, type='folder')
        #self.fe1.browse1_btn.configure(state='disabled')
        
        config_dict = ut.PSA_SND_read_config(const.strConfigFile)
        directory = f'{config_dict[const.strWorkingFolder]}{const.strDataResults}'
        filtered_dir = [identity for identity in os.listdir(directory) if len(identity)==20]
        sorted_dir = sorted(filtered_dir, key = lambda t: datetime.strptime(t[4:], '%Y-%m-%d-%H-%M'), reverse=True)

        self.fe1.dir_box.configure(state='normal')
        self.fe1.dir_box.delete("1.0", "end")
        self.fe1.dir_box.insert("1.0", f'{const.strDataResults}{sorted_dir[0]}')
        self.fe1.dir_box.configure(state='disabled')
        
        self.fe1.pack(padx=10, pady=10)
        # -----------------------------------------------------------
        '''Frame 1: Select Asset, Enter Description'''
        frame01 = ttk.LabelFrame(LabelFrame00, text='View Asset File')

        self.fe2 = ExcelEntry(frame01, type='file')
        self.bind('<FocusOut>', self.clear_canvas)
        
        self.fe2.pack(padx=10, pady=10)
 
        # -----------------------------------------------------------
        '''Frame 1: Select Asset, Enter Description'''
        frame1 = ttk.Frame(LabelFrame10)

        self.select_asset = tk.Label(frame1, text='Select Asset')
        self.select_asset.pack(padx=10, pady=10, anchor=tk.S)

        asset_list = ('No_Asset_File')
        self.pick_asset = ttk.Combobox(frame1, values=asset_list, textvariable=asset_value, width = 35, postcommand=lambda: self.pick_asset.configure(values=self.update_list()))
        self.pick_asset.current(0)
        self.pick_asset.bind('<<ComboboxSelected>>', self.refresh_image)
        self.pick_asset.pack(padx=10, pady=10, anchor=tk.N)

        self.sep = ttk.Separator(frame1, orient='horizontal')
        self.sep.pack(fill='x')

        self.threshold_label = tk.Label(frame1, text='Choose Threshold')
        self.threshold_label.pack(padx=10, pady=10, anchor=tk.S)

        self.threshold = tk.StringVar(frame1, '75%')
        self.threshold_input = tk.Entry(frame1, textvariable=self.threshold)
        self.threshold_input.bind('<Return>', self.refresh_image)
        self.threshold_input.pack(padx=10, pady=10, anchor=tk.N)

        frame1.pack()

        # -----------------------------------------------------------
        '''Grid Packing And Final Buttons for LabelFrame10'''
        self.cx = 900 # Width
        self.cy = 475 # Height

        #self.undo_btn = tk.Button(LabelFrame20, text='Undo', command=self.restore_file, bg='black', fg='white')
        #self.undo_btn.grid(pady=10, padx=10, ipadx=25, column=0, row=0, sticky=tk.NW)

        self.refresh_btn = tk.Button(LabelFrame20, text='Refresh Graph', command=self.refresh_image)
        self.refresh_btn.grid(pady=10, padx=10, column=1, row=0, sticky=tk.NSEW)
        
        self.view_thresholds_btn = tk.Button(LabelFrame20, text='View Thresholds', command=self.threshold_graph_btn_clicked, fg='blue')
        self.view_thresholds_btn.grid(pady=10, padx=10, column=0, row=0, sticky=tk.NSEW)
        
        self.plot_label = tk.Canvas(LabelFrame20, height=self.cy, width=self.cx, bg='white')
        self.plot_label.grid(pady=10, padx=10, column=0, row=1, sticky=tk.N, columnspan=4)

        self.save_btn = tk.Button(LabelFrame20, text='Save (Overwrite)', command=self.save, fg='red')
        self.save_btn.grid(pady=10, padx=10, column=2, row=0, sticky=tk.NSEW)

        self.saveas_btn = tk.Button(LabelFrame20, text='Save a Copy', command=self.save_as, fg='green')
        self.saveas_btn.grid(pady=10, padx=10, column=3, row=0, sticky=tk.NSEW)

        # -----------------------------------------------------------
        '''Final Packing'''       
        frame0.grid(column=0, row=0, columnspan=2, padx=10, pady=10)
        frame01.grid(column=0, row=1, columnspan=2, padx=10, pady=10)

        LabelFrame00.grid(padx=10, pady=10, column=0, row=0) # Update Outages
        LabelFrame10.grid(padx=10, pady=10, column=1, row=0) # Add New Outage
        LabelFrame20.grid(padx=10, pady=10, column=0, row=1, columnspan=2) # Plot Show

    # -------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    '''Functions for input processing'''
    def clear_canvas(self, event=None):
        if self.fe1.dir_box.get("1.0", "end").strip() == const.strNoFileDisp:
            self.plot_label.delete('all')
            self.pick_asset.configure(values=('No_Asset_File'))
            self.pick_asset.current(0)

    def threshold_graph_btn_clicked(self):
        status = const.PSAok
        asset_data = f'{self.dict_config[const.strWorkingFolder]}{self.fe2.dir_box.get("1.0","end").strip()}'
        tempfile = pa.plot_asset_data('', asset_data, relevant=False, save=True) # Change relevant to True to truncate results to where Events == True
        if os.path.isfile(tempfile):
            os.system('"' + tempfile + '"')
        else:
            status = const.PSAfileExistError
            msg = "Def threshold_graph_btn_clicked: File doesn't exist: " + tempfile
        if status != const.PSAok:
            messagebox.showwarning(title="WARNING", message=msg)

    def update_list(self):
        asset_list = ()

        if self.fe1.dir_box.get('1.0','end')[:-1] != const.strNoFileDisp and self.fe2.dir_box.get('1.0','end')[:-1] == const.strNoFileDisp:
            self.old_dir = self.fe1.dir_box.get('1.0','end')
            PSA_file = self.dict_config[const.strWorkingFolder] + self.fe1.dir_box.get('1.0', 'end')[:-1]
            PSA_file = PSA_file.replace('/','\\')
            PSARunID = PSA_file.split('\\')[-1]
            config_dict = ut.PSA_SND_read_runtimeparams(f'{PSA_file}/{PSARunID}{const.strRunTimeParams}')
            self.file = PSA_file + const.strINfolder + config_dict[const.strAssetData]
            self.file = self.file.replace('\\','/')
            filepath = self.file.replace(self.dict_config[const.strWorkingFolder],'')
            asset_list = tuple(sorted(pd.read_excel(self.file, usecols=[0, 1])['ASSET_ID']))     
        elif self.fe2.dir_box.get('1.0','end')[:-1] != const.strNoFileDisp:
            self.file = self.dict_config[const.strWorkingFolder] + self.fe2.dir_box.get('1.0','end')[:-1]
            self.file = self.file.replace('\\','/')
            filepath = self.file.replace(self.dict_config[const.strWorkingFolder],'')
            asset_list = tuple(sorted(pd.read_excel(self.file, usecols=[0, 1])['ASSET_ID']))

        if asset_list == ():
            self.save_btn.configure(fg='red')
            asset_list = ('No_Asset_File')
            self.fe2.editdata1_btn.configure(state='disabled')
        else: 
            if const.strDataInput in filepath.replace('/','\\'):
                self.save_btn.configure(fg='green')
                self.fe2.editdata1_btn.configure(state='normal')
            else:
                self.save_btn.configure(fg='red')
                self.fe2.editdata1_btn.configure(state='normal')
            self.fe2.dir_box.configure(state='normal')
            self.fe2.dir_box.delete('1.0','end')
            self.fe2.dir_box.insert('1.0',filepath)
            self.fe2.dir_box.configure(state='disabled')

        return asset_list

    '''
    def restore_file(self):
        if isinstance(self.df_backup, pd.DataFrame) and self.v > 0:
            self.df = self.df_backup
            print('DataFrame changed in memory')
    '''
    # -------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    '''Process Data 1'''
    '''
    def update_file(self):
        wb = openpyxl.load_workbook(filename=file)
        ws = wb.worksheets[0]
        ws_dict = {}

        for index, row in enumerate(ws.rows, start=1):
            ws_dict[row[0].value] = index
        
        primary = self.pick_asset.get()
        threshold = self.threshold.get()
        
        ws[f'C{ws_dict[primary]}'] = threshold
        wb.save(file)


        if not isinstance(self.df, pd.DataFrame):
            file = self.dict_config[const.strWorkingFolder] + self.fe2.dir_box.get('1.0', 'end')[:-1]
            self.df = pd.read_excel(file, index_col=0, sheet_name=[])
        
        self.df_backup = self.df.copy()
        primary = self.pick_asset.get()
        self.df.loc[primary,'MAX_LOADING'] = self.threshold.get()
        print('DataFrame changed in memory')
    '''

    def save(self):
        if self.fe1.dir_box.get('1.0','end') != const.strNoFileDisp:
            file = self.file
            if const.strDataInput not in file.replace('/','\\'):
                if not messagebox.askokcancel('Warning',f'This will overwrite historic data in:\n{file}\n\nAre you sure?'):
                    print('Save cancelled')
                    return
        
            wb = openpyxl.load_workbook(filename=file)
            ws = wb.worksheets[0]
            ws_dict = {}

            for index, row in enumerate(ws.rows, start=1):
                ws_dict[row[0].value] = index
            
            primary = self.pick_asset.get()
            if self.threshold_to_be_used == '':
                raise ValueError('No Threshold to input')
            threshold = float(self.threshold_to_be_used)
            try:
                float(threshold)
            except:
                threshold
            
            ws[f'C{ws_dict[primary]}'] = threshold # The column is always c, so we just need the row index for the primary, which we get using the dictionary
            wb.save(file)
            print(f'Saved to {file}')

    def save_as(self):
        if self.fe1.dir_box.get('1.0','end') != const.strNoFileDisp:
            file = self.file
            
            ct = datetime.now()
            file_name = f'{datetime.strftime(ct, "%d-%m-%y")} {const.strAssetData}_1.xlsx'
            save_file = f'{self.dict_config[const.strWorkingFolder]}{const.strDataInput}{file_name}'
            
            i = 1
            while os.path.isfile(save_file):
                save_file = f'{save_file[:-7]}_{i}.xlsx'
                i += 1

            if not messagebox.askokcancel('Info', f'This will generate a new file in:\n{save_file}'):
                print('Addition Cancelled')
                return

            wb = openpyxl.load_workbook(filename=file)
            ws = wb.worksheets[0]
            ws_dict = {}

            for index, row in enumerate(ws.rows, start=1):
                ws_dict[row[0].value] = index
            
            primary = self.pick_asset.get()
            if self.threshold_to_be_used == '':
                raise ValueError('No Threshold to input')
            threshold = float(self.threshold_to_be_used)

            ws[f'C{ws_dict[primary]}'] = threshold
            
            wb.save(save_file)
            self.file= save_file
            filepath = self.file.replace(self.dict_config[const.strWorkingFolder],'')
            messagebox.showinfo('Info',f'Now working from this file:\n{file}\n\nFurther changes can be safely saved without overwrite consequences')
            
            self.fe2.dir_box.configure(state='normal')
            self.fe2.dir_box.delete('1.0','end')
            self.fe2.dir_box.insert('1.0',filepath)
            self.fe2.dir_box.configure(state='disabled')
            self.save_btn.configure(fg='green')

            print(f'Saved to {save_file}')
    
    def refresh_image(self, event=None):
        primary = self.pick_asset.get().upper()
        BSP_file = f'{self.dict_config[const.strWorkingFolder]}/packages/psa/bsps_primary_txs_RATINGS_Thresholds_v1.xlsx'
        directory = self.dict_config[const.strWorkingFolder] + self.fe1.dir_box.get('1.0', 'end')[:-1]
        directory = directory.replace('/','\\')
        PSARunID = directory.split('\\')[-1]
        sheet = 'NET_DEM_MVA'
        try:
            threshold = float(self.threshold.get())
        except:
            threshold = self.threshold.get()
    
        print(f'Threshold: {threshold}')
        try:
            file, self.threshold_to_be_used = pa.plot_sia(primary, PSARunID, directory, sheet, threshold, decorate=True, save=True, BSP_file=BSP_file)
            self.plot = Image.open(file).resize((self.cx,self.cy))
            self.pic_plot = ImageTk.PhotoImage(self.plot)
            self.plot_label.create_image(self.cx/2,self.cy/2,image=self.pic_plot)
        except FileNotFoundError as err:
            messagebox.showinfo('Info', f'FileNotFoundError: No data for asset: {primary}')
            print(err)
        except BaseException as err:
            messagebox.showerror('Error', err)
            print(err)
        return 'refresh_image'

    def quit_me(self):
        print('Exiting')
        self.quit()
        self.destroy()

# -------------------------------------------------------------------------------------------------------------------------------------------------------------------------
if __name__ == '__main__':
    App().mainloop()