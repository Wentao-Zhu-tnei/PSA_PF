import os
import math
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
import PF_Config as pfconf
import UsefulPandas
import xlsxwriter
import openpyxl
import PSA_File_Validation as validation
import PSA_SND_Constants as const
import PSA_SND_Utilities as ut
import powerfactory_interface as pf
import csv
import time
import PSA_Events_Input_Output as eventsio
import PSA_Switch_Input_Output as switchio
import PSA_ProgressBar as pbar
from pathlib import Path
from collections import defaultdict

dict_config = ut.PSA_SND_read_config(const.strConfigFile)

def PF_getMasterRefList(pf, outputFile):
    dict_config = ut.PSA_SND_read_config(const.strConfigFile)

    objects = {
        "Line": "*.ElmLne",
        "2wdgTransformer": "*.ElmTr2",
        "Switch": "*.StaSwitch",
        "Coupler": "*.ElmCoup",
        "SyncMachine": "*.ElmSym"
    }

    lRef, lType = ([] for i in range(2))

    for key, value in objects.items():
        relevant_objects = pf.app.GetCalcRelevantObjects(value)
        for obj in relevant_objects:
            lRef.append(obj.loc_name)
            lType.append(key)

    df = pd.DataFrame([lRef, lType]).T
    df.columns = ["Asset", "Type"]
    df.to_excel(outputFile,header=True, index=False)
    copyFile = os.path.join(dict_config[const.strWorkingFolder], const.strRefAssetFile)
    df.to_excel(outputFile,header=True, index=False)
    df.to_excel(copyFile,header=True, index=False)
    return

def PF_readSwitchStatus(pf, output_folder):
    lSwitchCIM = []
    switches = pf.app.GetCalcRelevantObjects("*.StaSwitch")
    for switch in switches:
        strSwtich_cim_ID = str(switch.GetAttribute('e:cimRdfId'))
        lSwitchCIM.append(strSwtich_cim_ID)
    dfSwitchCIM = pd.DataFrame(data=lSwitchCIM, columns=['SwitchCIM'])
    dfSwitchCIM.to_excel(output_folder + "\\" + 'swtichCIM.xlsx')

def PF_calDownScaling(pf, dfAll2wTrafos, dfAllLoads, grid, output_folder, bDebug):
    print("################################################################")
    print(f"Calculating downscaling factor for grid: [{grid}]")
    print("################################################################")

    #### TRANSFORMERS ####
    if pfconf.bDebug:
        print('Getting ALL transformers including the PF objects as a column in the df')
    dfAll2wTrafos['Name_grid'] = dfAll2wTrafos['Name'] + '_' + dfAll2wTrafos['Grid']  ## this is the unique index proposed by AM
    #dfAll2wTrafos.to_excel(r"C:\Wentao Zhu\Project\PSA\Networks-Transition-Technical-Trials-PSA-1\data\Check\dfAll2wTrafos.xlsx")
    dfAll2wTrafos_grid = UsefulPandas.filter_df_single(dfAll2wTrafos, 'Grid', grid, regex=True)[
        ['Name', 'Grid', 'Name_grid',
         'From_LV', 'To_HV', 'S_rated',
         'S_rated_act', 'CIM_ID', 'PF_object']]
    if pfconf.bDebug:
        print(f'Number of transformers in [{grid}]: {dfAll2wTrafos_grid.shape[0]} out of {dfAll2wTrafos.shape[0]}')
    dfAll2wTrafos_grid_grp = dfAll2wTrafos_grid.groupby(['Grid', ]).agg({'Name': 'count', 'S_rated_act': np.sum, }). \
        sort_values(by=['Grid'], ascending=True)
    dictAll2wTrafos_grid_grp_kVA = dfAll2wTrafos_grid_grp.to_dict()['S_rated_act']
    dfAll2wTrafos_grid["Total_feeder_MVA"] = dfAll2wTrafos_grid["Grid"].map(dictAll2wTrafos_grid_grp_kVA)
    dfAll2wTrafos_grid['S_rated_indv_perc'] = dfAll2wTrafos_grid['S_rated_act'] * 100 / dfAll2wTrafos_grid[
        'Total_feeder_MVA']
    dfAll2wTrafos_grid_grp = dfAll2wTrafos_grid.groupby(['Grid', ]).agg({
        'Name': 'count',
        'S_rated_act': np.sum,
        'S_rated_indv_perc': np.sum}).sort_values(by=['Grid'], ascending=True)
    if pfconf.bDebug:
        print(f"{dfAll2wTrafos_grid_grp['S_rated_indv_perc']}")

    #### LOADS ####
    if pfconf.bDebug:
        print('Getting ALL loads including the PF objects as a column in the df')
    dfAllLoads['Name_grid'] = dfAllLoads['Name'] + '_' + dfAllLoads['Grid']  ## this is the unique index proposed by AM
    all_loads_df_grid = UsefulPandas.filter_df_single(dfAllLoads, 'Grid', grid, regex=True)[
        ['Name', 'Grid', 'Name_grid', 'Terminal', 'P_nom', 'Q_nom', 'CIM_ID', 'PF_object']]
    all_loads_df_grid = UsefulPandas.df_calc_kVA(all_loads_df_grid, 'S_nom', 'P_nom', 'Q_nom')
    if pfconf.bDebug:
        print(f'Number of loads in [{grid}]: {all_loads_df_grid.shape[0]} out of {dfAllLoads.shape[0]}')
        print(
            f'Number of transformers in [{grid}]: {dfAll2wTrafos_grid.shape[0]} and loads {all_loads_df_grid.shape[0]} equal: [{dfAll2wTrafos_grid.shape[0] == all_loads_df_grid.shape[0]}]')
    """
    calculate downscaling factors for load connected to secondary tx
    """
    ### If there is a load attached to the transformer
    dfAllTrafosLoads = pd.DataFrame(columns=['Name_transformer', 'S_rated_act', 'S_rated_indv_perc', 'Terminal',
                                             'Name_load', 'Grid', 'CIM_ID_transformer', 'CIM_ID_load',
                                             'PF_object_load'])

    ### If there is NO load attached to the transformer
    dfAllTrafos_grid_no_load = pd.DataFrame(columns=['Name_transformer', 'S_rated_act', 'S_rated_indv_perc',
                                                     'Terminal', 'Grid', 'CIM_ID_transformer'])

    if pfconf.bDebug:
        print('Get elements connected to the LV terminal of the transformers')
    for idx, Tx in dfAll2wTrafos_grid.iterrows():
        dfTerminalElements = pf.get_terminal_conn_elements(Tx['From_LV'])
        dfTerminalElements_load = UsefulPandas.filter_df_single(dfTerminalElements, 'Class', 'ElmLod', regex=False)
        if len(dfTerminalElements_load.index) > 0:
            dfAllTrafosLoads.loc[idx] = [Tx['Name'], Tx['S_rated_act'], Tx['S_rated_indv_perc'],
                                         Tx['From_LV'], dfTerminalElements_load['Name'].values[0],
                                         Tx['Grid'], Tx['CIM_ID'], dfTerminalElements_load['CIM_ID'].values[0],
                                         dfTerminalElements_load['PF_object'].values[0]]
        else:
            print('No lines/loads connected to this transformer')
            dfAllTrafos_grid_no_load.loc[idx] = [Tx['Name'], Tx['S_rated_act'], Tx['S_rated_indv_perc'], Tx['From_LV'],
                                                 Tx['Grid'], Tx['CIM_ID']]

    ### We have 2 transformers without a load and 1 transformer with both HV and LV terminals having the same name,
    # which causes an issue. RINA will update this later on.
    ### Merge this with the main load dataframe to get P_nom and Q_nom
    # print(dfAllTrafosLoads.columns)
    # print(all_loads_df_grid.columns)
    dfAllTrafos_grid_PQnom = pd.merge(dfAllTrafosLoads, all_loads_df_grid, on=['Terminal', 'Grid'], how='left',
                                      indicator=True)
    # print(dfAllTrafos_grid_PQnom.columns)
    dfDownScalingFactor_grid = dfAllTrafos_grid_PQnom[['Name_transformer', 'Name_load', 'Grid', 'Name_grid',
                                                       'S_rated_indv_perc', 'PF_object_load']]

    # if bDebug:
    if not os.path.isdir(output_folder):
        os.mkdir(output_folder)
    if pfconf.bDebug:
        print('----------------------------------------------------------------')
        print("Outputting downscaling factor")
        print('----------------------------------------------------------------')
    # dfDownScalingFactor_grid.to_excel(output_folder + 'PSA_DOWNSCALING_FACTORS' +'_' + str(grid) + '.xlsx')
    dfAllTrafos_grid_PQnom.to_excel(output_folder + 'PSA_DOWNSCALING_FACTORS_PQNOM' + '_' + str(grid) + '.xlsx')

    if pfconf.bDebug:
        print(dfAllTrafos_grid_PQnom[
                  '_merge'].unique())  # this is to verify that the number of loads aligns with the number of transformers

    return dfDownScalingFactor_grid, dfAllTrafos_grid_PQnom

def PF_combineOutages(bDebug, plout, unplout, file_plout, sheet_plout, dfUnplout, dir_combdoutages):
    """
    detect the type of outage input file and combine into one df
    """
    if plout:
        df_plout = validation.readMaintFile(file_plout, sheet_name=sheet_plout)
    else:
        df_plout = pd.DataFrame()
    if unplout:
        df_unplout = dfUnplout
    else:
        df_unplout = pd.DataFrame()

    frames = [df_plout, df_unplout]
    df_outages = pd.concat(frames)
    df_outages.reset_index(drop=True, inplace=True)
    if bDebug:
        if not os.path.isdir(dir_combdoutages):
            os.mkdir(dir_combdoutages)
        df_outages.to_excel(dir_combdoutages + "\\" + "COMBINED_OUTAGES.xlsx")

    return df_outages

def PF_mapOutages(strbUTC, start_time, df_outages, outageFolder, number_of_days, number_of_half_hours):
    """
    map outage periods to on/off status table of components
    """

    Time = pd.Series(
    ['%s:%s' % (h, m) for h in ([00] + list(range(1, int(number_of_half_hours / 2)))) for m in ('00', '30')])
    Days = pd.Series(['%s%s' % ('Day', d) for d in range(number_of_days)])
    
    df_outages = df_outages.reset_index(drop=True)

    lAssetIDs = df_outages.iloc[:]["ASSET_ID"]
    lAssetIDs = set(lAssetIDs)

    lDfOutages = []
    lOutageAssets = []
    lOutageAssetTypes = []

    if not os.path.isdir(outageFolder):
        os.mkdir(outageFolder)
    for asset_type in set(df_outages['ASSET_TYPE']):
        if not os.path.isdir(os.path.join(outageFolder,asset_type)):
            os.mkdir(os.path.join(outageFolder,asset_type))

    for AssetID in lAssetIDs:
        index_dfAsset = int(df_outages[df_outages['ASSET_ID'] == AssetID].index[0])
        AssetType = df_outages._get_value(index=index_dfAsset, col="ASSET_TYPE", takeable=False)
        lOutageAssets.append(AssetID)
        lOutageAssetTypes.append(AssetType)
        arrayStatus = np.zeros((number_of_days * number_of_half_hours,), dtype=int)  # outserv -> ZERO (in-service)
        for index in df_outages[df_outages['ASSET_ID'] == AssetID].index:
            deltaStartOutage = int((df_outages.loc[index]["START_OUTAGE"] - datetime.strptime(str(start_time),
                               "%Y-%m-%d %H:%M:%S")).total_seconds() / (60 * 30))

            deltaEndOutage = int((df_outages.loc[index]["END_OUTAGE"] - datetime.strptime(str(start_time),
                             "%Y-%m-%d %H:%M:%S")).total_seconds() / (60 * 30))
            if strbUTC == "TRUE":
                deltaStartOutage = deltaStartOutage - 2
                deltaEndOutage = deltaEndOutage - 2
            arrayStatus[(0 if deltaStartOutage<0 else deltaStartOutage):(0 if deltaEndOutage<0 else deltaEndOutage)] = 1  # outserv -> ONE (out-of-service)

        dfStatus = pd.DataFrame(1, index=np.arange(number_of_half_hours), columns=np.arange(number_of_days))
        for day in range(number_of_days):
            dfStatus.iloc[range(number_of_half_hours), day] = arrayStatus[day * number_of_half_hours: (day + 1) *
                                                              number_of_half_hours]
        lDfOutages.append(dfStatus)

        dfStatus.set_index(Time)
        asset_outage_file = AssetID + ".xlsx"
        dfStatus.to_excel(os.path.join(outageFolder,AssetType, asset_outage_file))

    return lOutageAssets, lOutageAssetTypes, lDfOutages

def PF_createXl(path_base, name, new_sheet_name):
    name = name.replace("/", "-")
    name = name.replace("\\", "-")
    workbook = xlsxwriter.Workbook(path_base + name + '.xlsx')
    worksheet = workbook.add_worksheet(new_sheet_name)
    workbook.close()

def PF_writeXl(dir_tgt, row, col, sheet_name, df):
    dir_tgt = dir_tgt.replace("\\", "/")
    if not os.path.isdir(dir_tgt):
        os.mkdir(dir_tgt)
    for file in os.listdir(dir_tgt):
        if file.endswith(".xlsx"):
            filename = file.replace(".xlsx", '')
            filename = filename.split('_Loading-')[1]
            wb = openpyxl.load_workbook(dir_tgt + "\\" + str(file), data_only=True)
            ws = wb[sheet_name]
            df_cell = df[df['Name'] == filename]['Loading'].iloc[0]
            ws.cell(row=row + 1, column=col + 1).value = float(df_cell)
            wb.save(dir_tgt + "\\" + str(file))

def PF_add_day_time(dir_tgt, sheet_name, number_of_days, number_of_half_hours):
    Time = pd.Series(
        ['%s:%s' % (h, m) for h in ([00] + list(range(1, int(number_of_half_hours / 2)))) for m in ('00', '30')])
    Days = pd.Series(['%s%s' % ('Day', d) for d in range(number_of_days)])
    
    for file in os.listdir(dir_tgt):
        if file.endswith(".xlsx"):
            wb = openpyxl.load_workbook(dir_tgt + "\\" + str(file), data_only=True)
            ws = wb[sheet_name]
            ws.insert_rows(idx=1, amount=1)
            ws.insert_cols(idx=1, amount=1)
            for i in range(2, 50):
                ws.cell(row=i, column=1).value = Time[i - 2]
            for j in range(2, 13):
                ws.cell(row=1, column=j).value = Days[j - 2]
            wb.save(dir_tgt + "\\" + str(file))

def PF_Write_SIA_Load(target_folder):
    dictFeederLoads = dict()
    for r, d, f in os.walk(target_folder):
        for file in f:
            if "SIA_Feeder" in file:
                strFeederName = file.replace("SIA_Feeder", "")
                strFeederName = strFeederName.replace(".xlsx", "")
                SIA_load = pd.read_excel(target_folder + "\\" + file, sheet_name="UND_DEM_MVA", index_col=0)
                dictFeederLoads[strFeederName] = SIA_load
    lFeeders = dictFeederLoads.keys()
    fSIALoadMVA = target_folder + "\\" + 'SIA_Load_MVA.xlsx'
    writer = pd.ExcelWriter(fSIALoadMVA, engine='xlsxwriter')
    for key, value in dictFeederLoads.items():
        value.to_excel(writer, sheet_name=key)
    # writer.save()
    return lFeeders, fSIALoadMVA, dictFeederLoads

def PF_Write_SIA_Gen(target_folder):
    dictGens = dict()
    for r, d, f in os.walk(target_folder):
        for file in f:
            if "SIA_Generator_" in file:
                strGenName = file.replace("SIA_Generator_", "")
                strGenName = strGenName.replace(".xlsx", "")
                SIA_gen = pd.read_excel(target_folder + "\\" + file, sheet_name="GEN_MW", index_col=0)
                dictGens[strGenName] = SIA_gen
    lGens = dictGens.keys()
    fSIAGenMW = target_folder + "\\" + 'SIA_Generator_MW.xlsx'
    writer = pd.ExcelWriter(fSIAGenMW, engine='xlsxwriter')
    for key, value in dictGens.items():
        value.to_excel(writer, sheet_name=key)
    # writer.save()
    return lGens, fSIAGenMW, dictGens

def PF_get_outages(lLines, lTrafos_2w, lOutageAssets):
    dictOutagesLine, dictOutagesTrafo = (dict() for i in range(2))
    for i in range(len(lOutageAssets)):
        for oLine in lLines:
            if oLine.loc_name == lOutageAssets[i]:
                dictOutagesLine[oLine.loc_name] = oLine.outserv

        for oTrafo_2w in lTrafos_2w:
            if oTrafo_2w.loc_name == lOutageAssets[i]:
                dictOutagesTrafo[oTrafo_2w.loc_name] = oTrafo_2w.outserv

    return dictOutagesLine, dictOutagesTrafo

def PF_set_outages(pf_object, dictOutagesLine, dictOutagesTrafo):
    for line, lineoutage in dictOutagesLine.items():
        pf_object.app.GetCalcRelevantObjects(line + ".ElmLne")[0].outserv = lineoutage
    for trafo, trafooutage in dictOutagesTrafo.items():
        pf_object.app.GetCalcRelevantObjects(trafo + ".ElmLne")[0].outserv = trafooutage
    return

def PF_update_Inom(lLines, lTrafos, dictMonitoredLineAssets, dictMonitoredTrafoAssets, lne_thresh, trafo_thresh):
    # Update Inom of transformers and lines, and return a dataframe with the name of the assets and the Inom_old and
    # Inom new of the assets (PF_Project not changed)
    lLineData = []
    lTrafoData = []
    for oLine in lLines:
        if oLine.loc_name in dictMonitoredLineAssets.keys():
            Inom_new_lne = dictMonitoredLineAssets[oLine.loc_name] * oLine.Inom_a
        else:
            Inom_new_lne = lne_thresh * oLine.Inom_a / 100
        lLineData.append([oLine.loc_name, oLine.Inom_a, Inom_new_lne])

    dfLine = pd.DataFrame(data=lLineData, columns=["asset_name", "Inom_old", "Inom_new"])

    for oTrafo in lTrafos:
        if oTrafo.loc_name in dictMonitoredTrafoAssets.keys():
            Inom_new_trafo = dictMonitoredTrafoAssets[oTrafo.loc_name] * oTrafo.Inom_h_a
        else:
            Inom_new_trafo = trafo_thresh * oTrafo.Inom_h_a / 100
        lTrafoData.append([oTrafo.loc_name, oTrafo.Inom_h_a, Inom_new_trafo])

    dfTrafo = pd.DataFrame(data=lTrafoData, columns=["asset_name", "Inom_old", "Inom_new"])

    return dfLine, dfTrafo

def PF_update_dtformat(strInput):
    """
    Update the format of the date time from str '%Y-%m-%d-%H-%M' to dt '%Y-%m-%d %H:%M:%S'
    """
    dtOutput = datetime.strptime(strInput, '%Y-%m-%d-%H-%M')
    dtOutput = dtOutput.strftime("%Y-%m-%d %H:%M")
    dtOutput = datetime.strptime(dtOutput, '%Y-%m-%d %H:%M')

    return dtOutput

def PF_read_cont(start_outage, end_outage, PSArunID):
    if isinstance(start_outage, str):
        start_outage = datetime.strptime(start_outage, "%Y-%m-%d %H:%M:%S")
    if isinstance(end_outage, str):
        end_outage = datetime.strptime(end_outage, "%Y-%m-%d %H:%M:%S")

    current_datetime = PF_update_dtformat(PSArunID[4:])

    if current_datetime <= end_outage and current_datetime >= start_outage:
        isOutage = True
    else:
        isOutage = False

    return isOutage

def PF_update_cont(CONT_file, PSArunID, number_of_days):
    cont_data = validation.readContFile(CONT_file,sheet_name=const.strDataSheet)
    for index, row in cont_data.iterrows():
        isOutage = PF_read_cont(row['START_OUTAGE'], row['END_OUTAGE'], PSArunID)
        current_datetime = PF_update_dtformat(PSArunID[4:])
        if isOutage:
            end_datetime = current_datetime + timedelta(days=number_of_days + 2)
            row['END_OUTAGE'] = end_datetime.strftime("%Y-%m-%d %H:%M") + ":00"
            cont_data.loc[index, "END_OUTAGE"] = row['END_OUTAGE']
        else:
            row['START_OUTAGE'] = (current_datetime + timedelta(-2)).strftime("%Y-%m-%d %H:%M") + ":00"
            row['END_OUTAGE'] = row['START_OUTAGE']
            cont_data.loc[index, "START_OUTAGE"] = row['START_OUTAGE']
            cont_data.loc[index, "END_OUTAGE"] = row['END_OUTAGE']
    # cont_data.to_excel(CONT_file, sheet_name="DATA", index=False)

    return cont_data

def PF_update_outages(pf, dictDfLine, dictDfTrafos_2w, dictDfSynm, dictDfLoad, i_half_hour, i_day, verbose=False):
    if bool(dictDfLine):
        for Line, dfLineOutages in dictDfLine.items():
            oLine = pf.app.GetCalcRelevantObjects(Line + ".ElmLne")[0]
            oLine.outserv = int(dfLineOutages.iloc[i_half_hour, i_day])
            if verbose and oLine.outserv:
                print(f'{oLine.loc_name} is outserv: {bool(oLine.outserv)}')
    if bool(dictDfTrafos_2w):
        for Trafo, dfTraoOutages in dictDfTrafos_2w.items():
            oTrafo = pf.app.GetCalcRelevantObjects(Trafo + ".ElmTr2")[0]
            oTrafo.outserv = int(dfTraoOutages.iloc[i_half_hour, i_day])
            if verbose and oTrafo.outserv:
                print(f'{oTrafo.loc_name} is outserv: {bool(oTrafo.outserv)}')
    if bool(dictDfSynm):
        for Synm, dfSynmOutages in dictDfSynm.items():
            oSynm = pf.app.GetCalcRelevantObjects(Synm + ".ElmSym")[0]
            oSynm.outserv = int(dfSynmOutages.iloc[i_half_hour, i_day])
            if verbose and oSynm.outserv:
                print(f'{oSynm.loc_name} is outserv: {bool(oSynm.outserv)}')
    if bool(dictDfLoad):
        for Load, dfLoadOutages in dictDfLoad.items():
            oLoad = pf.app.GetCalcRelevantObjects(Load + ".ElmLod")[0]
            print(dfLoadOutages.iloc[i_half_hour, i_day])
            oLoad.outserv = int(dfLoadOutages.iloc[i_half_hour, i_day])
            if verbose and oLoad.outserv:
                print(f'{oLoad.loc_name} is outserv: {bool(oLoad.outserv)}')
    return

def getSign(P):
    if P > 0:
        sign = 1
    else:
        sign = -1
    return sign

def findcolrow(dt, startdate):
    dt = str(dt)
    date = dt.split(" ")[0]
    time = dt.split(" ")[1]
    dt_date = datetime.strptime(str(date),"%Y-%m-%d")
    startdate = datetime.strptime(str(startdate),"%Y-%m-%d")
    starttime = datetime.strptime(str("00:00:00"),"%H:%M:%S")
    dt_time = datetime.strptime(str(time),"%H:%M:%S")
    delta_days = (dt_date - startdate).days
    delta_half_hours = int(divmod((dt_time - starttime).total_seconds(), 1800)[0])
    return delta_days, delta_half_hours

def PF_CalcAssetLoading(isAuto, PSArunID, bEvents, dfEvents, bSwitch, dfSwitch, plout, unplout, target_NeRDA_folder, pf, no_days, no_half_hours,
                        lMonitoredLineAssets, lMonitoredTrafoAssets, lOutageAssets, lOutageAssetTypes, lDfOutages, dictSIAFeeder,
                        dictSIAGen, output_folder, power_factor, lagging, start_time_row, bDebug, dir_lnelding,
                        dir_trafolding, PSA_running_mode, lNamesGrid, dfDownScalingFactor_grid, dfAllTrafos_grid_PQnom, PSAfolder,
                        number_of_days, number_of_half_hours):
    """
    update the connectivity status of components (in-/out-of-service)
    current_time: the number of half-hour intervals from start
    """
    Time = pd.Series(
        ['%s:%s' % (h, m) for h in ([00] + list(range(1, int(number_of_half_hours / 2)))) for m in ('00', '30')])
    Days = pd.Series(['%s%s' % ('Day', d) for d in range(number_of_days)])

    print("\n\n################################################################")
    print(f"NOW RUNNING IN {PSA_running_mode} MODE\n")
    print("################################################################\n\n")
    # Record performance time stamp
    t0 = time.time()
    status = const.PSAok
    msg = ""
    power_factor = float(power_factor)
    power_factor2 = math.sin(math.acos(power_factor))
    lLines = pf.app.GetCalcRelevantObjects("*.ElmLne")
    lTrafos_2w = pf.app.GetCalcRelevantObjects('*.ElmTr2')
    lSynm = pf.app.GetCalcRelevantObjects('*.ElmSym')
    lLoad = pf.app.GetCalcRelevantObjects('*.ElmLod')
    # dfAll2wTrafos = pf.get_transformers_2w_results()
    dfAll2wTrafos = pf.get_transformers_2w_results()
    dfAllLoads = pf.get_loads_results()
    dictLineLoading, dictTrafoLoading, dictLinePowerBus1, dictLineIBus1, dictLineIBus2,\
        dictTrafoPowerLv, dictTrafoIBus1, dictTrafoIBus2 = (dict() for i in range(8))
    scenario_folder = Path(output_folder)
    
    if bEvents:
        dictDfServices = eventsio.mapServices(strbUTC=dict_config[const.bUTC],start_time=PSArunID[4:14]+" "+"00:00:00",
                                            dfEvents=dfEvents, PSAfolder=PSAfolder,
                                            number_of_days=number_of_days, number_of_half_hours=number_of_half_hours)

    if bSwitch:
        dictSwitchStatus = switchio.mapSwitchStatus(strbUTC=dict_config[const.bUTC],start_time=PSArunID[4:14]+" "+"00:00:00",
                                                    dfSwitch=dfSwitch, PSAfolder=PSAfolder,number_of_days=number_of_days,
                                                    number_of_half_hours=number_of_half_hours)

    if PSA_running_mode == "M":
        scenario = "MAINT"
        dictDfTrafos_2w, dictDfLine, dictDfSynm, dictDfLoad = ut.readOutageFile(scenario, PSAfolder)
    elif PSA_running_mode == "C":
        scenario = "CONT"
        dictDfTrafos_2w, dictDfLine, dictDfSynm, dictDfLoad = ut.readOutageFile(scenario, PSAfolder)
    elif PSA_running_mode == "MC":
        scenario = "MAINT_CONT"
        dictDfTrafos_2w, dictDfLine, dictDfSynm, dictDfLoad = ut.readOutageFile(scenario, PSAfolder)
    elif PSA_running_mode == "B":
        scenario = "BASE"
    else:
        print("PF_Calc_assetLoading INVALID PSA_running_mode = " + PSA_running_mode)


    if bDebug:
        if not os.path.isdir(dir_lnelding + "\\" + "PSA_" + PSA_running_mode + "_LINE_LOADING"):
            os.mkdir(dir_lnelding + "\\" + "PSA_" + PSA_running_mode + "_LINE_LOADING")
        if not os.path.isdir(dir_trafolding + "\\" + "PSA_" + PSA_running_mode + "_TRANSFORMER_LOADING"):
            os.mkdir(dir_trafolding + "\\" + "PSA_" + PSA_running_mode + "_TRANSFORMER_LOADING")
    for strLineAsset in lMonitoredLineAssets:
        dictLineLoading[strLineAsset], dictLinePowerBus1[strLineAsset], dictLineIBus1[strLineAsset], \
        dictLineIBus2[strLineAsset],  = (pd.DataFrame(np.nan, index=[i for i in range(number_of_half_hours)],
                                                     columns=[i for i in range(number_of_days)]) for i in range(4))
    for strTrafoAsset in lMonitoredTrafoAssets:
        dictTrafoLoading[strTrafoAsset], dictTrafoPowerLv[strTrafoAsset], dictTrafoIBus1[strTrafoAsset], \
        dictTrafoIBus2[strTrafoAsset]  = (pd.DataFrame(np.nan, index=[i for i in range(number_of_half_hours)],
                                                       columns=[i for i in range(number_of_days)]) for i in range(4))
    if bDebug:
        lAttr = ['b:loc_name', 'e:cpGrid', 'e:cimRdfId', 'c:loading', 'e:outserv', 'm:P:bus1', 'm:I:bus1', 'm:I:bus2']
        lHeader = ['Name', 'Grid', 'CIM_ID', 'Loading', 'IN/OUT_of_service', 'bus1/buslv power', 'bus1/bushlv current',
                    'bus2/bushv current']
    else:
        lAttr = ['b:loc_name', 'c:loading', 'm:P:bus1', 'm:I:bus1', 'm:I:bus2']
        lHeader = ['Name', 'Loading', 'bus1/buslv power', 'bus1/bushlv current', 'bus2/bushv current']

    lDownScalingFactor_grid = []
    lAllTrafos_grid_PQnom = []

    if bDebug:
        dictInService_line, dictInService_trafo, dictInService_synm, dictInService_load = (dict() for i in range(4))
        for oLine in lLines:
            strLineAsset = oLine.loc_name
            dictInService_line[strLineAsset] = pd.DataFrame(np.nan,index=[i for i in range(number_of_half_hours)],
                                                            columns=[i for i in range(number_of_days)])
        for oTrafo_2w in lTrafos_2w:
            strTrafoAsset = oTrafo_2w.loc_name
            dictInService_trafo[strTrafoAsset] = pd.DataFrame(np.nan,index=[i for i in range(number_of_half_hours)],
                                                              columns=[i for i in range(number_of_days)])
        for oSynm in lSynm:
            strSynmAsset = oSynm.loc_name
            dictInService_synm[strSynmAsset] = pd.DataFrame(np.nan,index=[i for i in range(number_of_half_hours)],
                                                              columns=[i for i in range(number_of_days)])
        for oLoad in lLoad:
            strLoadAsset = oLoad.loc_name
            dictInService_load[strLoadAsset] = pd.DataFrame(np.nan,index=[i for i in range(number_of_half_hours)],
                                                              columns=[i for i in range(number_of_days)])
    # Record performance time stamp
    header = ["DAY", "HH", "RUN_INIT", "CUM_TIME", "(UN)PLOUT", "SIA_GENS", "DS_FACTOR", "FEEDERS", "LF_CALC",
              "INSVC_LINE", "RESULTS", "TIMESTAMP", "FEED-60", "FEED-61", "FEED-62", "C1", "MEMORY(%)"]
    perf_file = os.path.dirname(os.path.dirname(output_folder)) + '/' + PSA_running_mode + '_performance_log.csv'
    with open(perf_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        f.close()
    t1 = time.time()

    ### Transform dictSIAFeeder into 3D dataframe ###
    npTime = np.tile(Time, len(Days))
    npDays = np.repeat(Days, len(Time))
    ldfSIAFeeders_multidx = []
    for strNameFeeder, dictFeeder in dictSIAFeeder.items():
        dfSIAFeeder = dictFeeder['UND_DEM_MVA']
        npSIAFeeder = np.reshape(np.transpose(dfSIAFeeder.to_numpy()),
                                 (1, number_of_days * number_of_half_hours))
        dfSIAFeeder_multidx = pd.DataFrame(data=npSIAFeeder, columns=pd.MultiIndex.from_tuples(zip(npDays, npTime)),
                                           index=[strNameFeeder])
        ldfSIAFeeders_multidx.append(dfSIAFeeder_multidx)
    dfSIAFeeders_multidx = pd.concat(ldfSIAFeeders_multidx)

    for i_day in range(no_days):
        if i_day == 0:
            if True:  # TODO: Global variable dfDownScalingFactor_new doesn't like long names (pfd data objects)
                lObjloads = []
                print("----------------------------------------------------------------")
                print("Calculating Downscaling Factor")
                print("----------------------------------------------------------------")

                pb=20

                for NameGrid in lNamesGrid:
                    if NameGrid != "Summary Grid":
                        pb=pb+1
                        pbar.updateProgressBar(pb, "Calculating Downscaling Factors for " + NameGrid)
                        dfDownScalingFactor_grid, dfAllTrafos_grid_PQnom = PF_calDownScaling(pf, dfAll2wTrafos,
                                                                           dfAllLoads, NameGrid, output_folder, bDebug)
                        lDownScalingFactor_grid.append(dfDownScalingFactor_grid)
                        lAllTrafos_grid_PQnom.append(dfAllTrafos_grid_PQnom)
                pbs = pb
                dfDownScalingFactor_grid = pd.concat(lDownScalingFactor_grid)
                dfAllTrafos_grid_PQnom = pd.concat(lAllTrafos_grid_PQnom)
                dfAllTrafos_grid_PQnom.to_excel(output_folder + 'ALL_PSA_DOWNSCALING_FACTORS_PQNOM.xlsx')
                # dfAllTrafos_grid_PQnom.to_pickle(output_folder + 'ALL_PSA_DOWNSCALING_FACTORS_PQNOM.pkl')
                # for index, row in dfDownScalingFactor_grid.iterrows():
                # global dfDownScalingFactor_new
                dfDownScalingFactor_new = dfDownScalingFactor_grid.drop(labels="Name_transformer", axis=1)
                # dfDownScalingFactor_new.to_pickle(output_folder + 'dfDownScalingFactor_new_pfobj.pkl')
                dfDownScalingFactor_new_csv = dfDownScalingFactor_new.copy(deep=True)
                dfDownScalingFactor_new_csv["PF_object_load"] = dfDownScalingFactor_new_csv["PF_object_load"].apply(str)
                dfDownScalingFactor_new_csv.to_csv(output_folder + 'dfDownScalingFactor_new_pfobj.csv')
                del dfDownScalingFactor_new_csv
                print("----------------------------------------------------------------")
                print("Downscaling Factor Calculation Completed")
                print("----------------------------------------------------------------")
            status = const.PSAok
            msg = ""
            range_half_hours = range(start_time_row, no_half_hours)

        else:
            range_half_hours = range(no_half_hours)
        print("----------------------------------------------------------------")
        print(f"Power flow calculation started for day {i_day}...")
        print("----------------------------------------------------------------")

        #disable PF db updates
        pf.enableLocalCache()

        for i_half_hour in range_half_hours:

            Hour = datetime.strptime("00:00", '%H:%M') + timedelta(hours=i_half_hour * 0.5)
            pct = float((i_day * no_half_hours) + i_half_hour)/float((no_days * no_half_hours))
            pb = pbs + int(float(90-pbs) * pct)
            pstr = "Processing: " + scenario + " Day " + str(i_day + 1) + " of " + str(no_days) + " - Time " + str(Hour)[11:16]
            pbar.updateProgressBar(pb, pstr)

            time_start = datetime.now()
            # Record performance time stamp
            t2 = time.time()


            # update line and trafo outages
            if bEvents:
                eventsio.updateServices(pf,dictDfServices,i_half_hour, i_day, verbose=True)

            if bSwitch:
                switchio.updateSwitchStatus(pf, dictSwitchStatus, i_half_hour, i_day, verbose=True)

            if scenario == "MAINT" or scenario == "CONT" or scenario == "MAINT_CONT":
                PF_update_outages(pf, dictDfLine, dictDfTrafos_2w, dictDfSynm, dictDfLoad, i_half_hour, i_day,
                                  verbose=False)
        
            # Record performance time stamp
            t3 = time.time()
            for strNameGen, dictGen in dictSIAGen.items():
                dfSIAGen = abs(dictGen['GEN_MW'])
                oGen = pf.app.GetCalcRelevantObjects(strNameGen + '.ElmSym')[0]
                pgen = float(dfSIAGen.iloc[i_half_hour, i_day])
                oGen.pgini = float(pgen)

            # Record performance time stamp
            t4 = time.time()

            # Initialise lists
            lP_load, lQ_load, lS_load, lName_load, lName_loadgrid = ([] for i in range(5))
            
            #df_test = dfDownScalingFactor_new.to_dict('records')

            # Record performance time stamp
            t5 = time.time()

            ### Used in loop below ###
            day = "Day" + str(i_day)
            hh = Time[i_half_hour]

            ### Iterate through dfDownScalingFactor_new ###
            c1 = 0

            # for index, row in dfDownScalingFactor_new.iterrows():
            for oLoad, feeder_name, load_name in zip(dfDownScalingFactor_new['PF_object_load'],
                                         dfDownScalingFactor_new['Grid'], dfDownScalingFactor_new['Name_load']):
                #memory = psutil.virtual_memory().percent
                memory = -999
                t50 = time.time()
                c1 = c1 + 1
                t60 = time.time()

                t51 = time.time()
                feeder_load = dfSIAFeeders_multidx.loc[feeder_name].loc[day].loc[hh]
                dfDownScalingFactor_intm = dfDownScalingFactor_new[dfDownScalingFactor_new['Grid'] == feeder_name]
                dfDownScalingFactor_intm_1 = dfDownScalingFactor_intm[dfDownScalingFactor_intm['Name_load'] == load_name]
                downscalingfactor = dfDownScalingFactor_intm_1['S_rated_indv_perc']
                df = downscalingfactor.iloc[0]
                t61 = time.time()
                
                t52 = time.time()
                Slini = feeder_load * df / 100
                Plini = Slini * power_factor
                if lagging:
                    Qlini = Slini * power_factor2
                else:
                    Qlini = -Slini * power_factor2
                oLoad.mode_inp = "SC"
                oLoad.slini = Slini
                oLoad.plini = Plini
                oLoad.qlini = Qlini
                t62 = time.time()

                if bDebug:
                    lName_loadgrid.append(feeder_name)
                    lS_load.append(Slini)
                    lP_load.append(Plini)
                    lQ_load.append(Qlini)
                    lName_load.append(load_name)

            if bDebug:
                df_load = pd.DataFrame(list(zip(lS_load, lP_load, lQ_load, lName_loadgrid)),
                                       columns=['S', 'P', 'Q', 'Grid'], index=lName_load)
                feeder_load_folder = os.path.dirname(os.path.dirname(output_folder)) + "\\" + "FEEDER_LOAD"
                if not os.path.isdir(feeder_load_folder):
                    os.mkdir(feeder_load_folder)
                df_load.to_excel(
                    feeder_load_folder + "\\" + "Load" + "_" + str(i_day) + "-" + str(i_half_hour) + ".xlsx")

            # Record performance time stamp
            t6 = time.time()
            # print(f"updating loads for one iteration takes {t6 - t5} seconds")

            #pf.disableLocalCache()
            pf.updatePFdb()
            ierr = pf.run_loadflow()
            #pf.enableLocalCache()

            # Record performance time stamp
            t7 = time.time()

            if ierr == 0:
                status = const.PSAok
                Hour = datetime.strptime("00:00", '%H:%M') + timedelta(hours=i_half_hour * 0.5)
                if bDebug:
                    for oLine in lLines:
                        dictInService_line[oLine.loc_name].iloc[i_half_hour, i_day] = oLine.outserv
                    for oTrafo_2w in lTrafos_2w:
                        dictInService_trafo[oTrafo_2w.loc_name].iloc[i_half_hour, i_day] = oTrafo_2w.outserv
                    for oSynm in lSynm:
                        dictInService_synm[oSynm.loc_name].iloc[i_half_hour, i_day] = oSynm.outserv
                    for oLoad in lLoad:
                        dictInService_load[oLoad.loc_name].iloc[i_half_hour, i_day] = oLoad.outserv
            else:
                if isAuto:
                    continue
                else:
                    pf.app.Show()
                    status = const.PSAPFError
                    msg = "Load flow error in {} day and {} half hour".format(i_day, i_half_hour)
                    return status, msg, dictLineLoading, dictTrafoLoading, dfDownScalingFactor_grid, \
                           dfAllTrafos_grid_PQnom, dictLinePowerBus1, dictLineIBus1, dictLineIBus2, dictTrafoPowerLv, \
                           dictTrafoIBus1, dictTrafoIBus2
            # Record performance time stamp
            t8 = time.time()
            if status == const.PSAok:
                df_line = pf.get_asset_results(type=const.strAssetLine, bAllAssets=False, lSubAssetsNames=lMonitoredLineAssets,
                                               lAttr=lAttr, lHeader=lHeader)
                for strLineAsset, dfLineAsset in dictLineLoading.items():
                    if df_line.loc[df_line['Name'] == strLineAsset]['Loading'].empty:
                        dfLineAsset.iloc[i_half_hour][i_day] = 0
                    else:
                        dfLineAsset.iloc[i_half_hour][i_day] = float(
                            df_line.loc[df_line['Name'] == strLineAsset]['Loading'].iloc[0])
                for strLineAsset, dfLineAsset in dictLinePowerBus1.items():
                    if df_line.loc[df_line['Name'] == strLineAsset]['bus1/buslv power'].empty:
                        dfLineAsset.iloc[i_half_hour][i_day] = 0
                    else:
                        dfLineAsset.iloc[i_half_hour][i_day] = float(
                            df_line.loc[df_line['Name'] == strLineAsset]['bus1/buslv power'].iloc[0])
                for strLineAsset, dfLineAsset in dictLineIBus1.items():
                    if df_line.loc[df_line['Name'] == strLineAsset]['bus1/bushlv current'].empty:
                        dfLineAsset.iloc[i_half_hour][i_day] = 0
                    else:
                        dfLineAsset.iloc[i_half_hour][i_day] = float(
                            df_line.loc[df_line['Name'] == strLineAsset]['bus1/bushlv current'].iloc[0])
                for strLineAsset, dfLineAsset in dictLineIBus2.items():
                    if df_line.loc[df_line['Name'] == strLineAsset]['bus2/bushv current'].empty:
                        dfLineAsset.iloc[i_half_hour][i_day] = 0
                    else:
                        dfLineAsset.iloc[i_half_hour][i_day] = float(
                            df_line.loc[df_line['Name'] == strLineAsset]['bus2/bushv current'].iloc[0])
                # for strLineAsset,dfLineAsset in dictLineLoading.items():
                #     lLineLoading.append(float(df_line.loc[df_line['Name'] == strLineAsset]['Loading'].iloc[0]))
                #     x = float(df_line.loc[df_line['Name'] == strLineAsset]['Loading'].iloc[0])
                #     dfLineAsset.iloc[i_half_hour][i_day] = x

                if bDebug:
                    df_line.to_excel(dir_lnelding + "\\" + "PSA_" + PSA_running_mode + "_LINE_LOADING" + "\\" + str(
                        i_day) + "d" + "-" + str(i_half_hour) + "hr" + '.xlsx', sheet_name="DATA")

                df_trafo = pf.get_asset_results(type=const.strAssetTrans, bAllAssets=False, lSubAssetsNames=lMonitoredTrafoAssets,
                                                lAttr=lAttr, lHeader=lHeader)
                for strTrafoAsset, dfTrafoAsset in dictTrafoLoading.items():
                    if df_trafo.loc[df_trafo['Name'] == strTrafoAsset]['Loading'].empty:
                        dfTrafoAsset.iloc[i_half_hour][i_day] = 0
                    else:
                        dfTrafoAsset.iloc[i_half_hour][i_day] = float(
                            df_trafo.loc[df_trafo['Name'] == strTrafoAsset]['Loading'].iloc[0])
                for strTrafoAsset, dfTrafoAsset in dictTrafoPowerLv.items():
                    if df_trafo.loc[df_trafo['Name'] == strTrafoAsset]['bus1/buslv power'].empty:
                        dfTrafoAsset.iloc[i_half_hour][i_day] = 0
                    else:
                        dfTrafoAsset.iloc[i_half_hour][i_day] = float(
                            df_trafo.loc[df_trafo['Name'] == strTrafoAsset]['bus1/buslv power'].iloc[0])
                for strTrafoAsset, dfTrafoAsset in dictTrafoIBus1.items():
                    if df_trafo.loc[df_trafo['Name'] == strTrafoAsset]['bus1/bushlv current'].empty:
                        dfTrafoAsset.iloc[i_half_hour][i_day] = 0
                    else:
                        dfTrafoAsset.iloc[i_half_hour][i_day] = float(
                            df_trafo.loc[df_trafo['Name'] == strTrafoAsset]['bus1/bushlv current'].iloc[0])
                for strTrafoAsset, dfTrafoAsset in dictTrafoIBus2.items():
                    if df_trafo.loc[df_trafo['Name'] == strTrafoAsset]['bus2/bushv current'].empty:
                        dfTrafoAsset.iloc[i_half_hour][i_day] = 0
                    else:
                        dfTrafoAsset.iloc[i_half_hour][i_day] = float(
                            df_trafo.loc[df_trafo['Name'] == strTrafoAsset]['bus2/bushv current'].iloc[0])
                if bDebug:
                    df_trafo.to_excel(
                        dir_trafolding + "\\" + "PSA_" + PSA_running_mode + "_TRANSFORMER_LOADING" + "\\" + str(
                            i_day) + "d" + "-" + str(i_half_hour) + "hr" + '.xlsx', sheet_name="DATA")

            time_end = datetime.now()
            time_elapsed = time_end - time_start

            # print("Elapsed time: " + str(time_elapsed))
            print(f"Day: {i_day}, Hour: {str(Hour)[11:16]}, Elapsed time (MM:SS): {str(time_elapsed)} ldf_ierr:{ierr}")
            # Record performance time stamp
            t9 = time.time()

            with open(perf_file, 'a', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(
                    [str(i_day), str(i_half_hour), str(t1 - t0), str(t2 - t1), str(t3 - t2), str(t4 - t3), str(t5 - t4),
                     str(t6 - t5), str(t7 - t6), str(t8 - t7), str(t9 - t8), datetime.now(), str(t60 - t50),
                     str(t61 - t51), str(t62 - t52), str(c1), str(memory)])
                f.close()
        
        #enable write to DB after end of HH loop
        pf.disableLocalCache()

    if bDebug:
        print("################################################################")
        print("Outputting lines, transformers, synchronous machine and load status data...")
        print("################################################################")
        pb=pb+1
        pbar.updateProgressBar(pb+1, "Debug Mode: Outputting Line Loading Data Files")
        for strLineAsset, dfInService_line in dictInService_line.items():
            strLineAsset = strLineAsset.replace("/", "-")
            strLineAsset = strLineAsset.replace("\\", "-")
            line_status_folder = target_NeRDA_folder + "\\" + "Line_Status"
            if not os.path.isdir(line_status_folder):
                os.mkdir(line_status_folder)
            dfInService_line.to_excel(line_status_folder + "\\" + strLineAsset + ".xlsx")
        pb=pb+1
        pbar.updateProgressBar(pb+1, "Debug Mode: Outputting Transformer Loading Data Files")
        print("Lines completed...")
        for strTrafoAsset, dfInService_trafo in dictInService_trafo.items():
            strTrafoAsset = strTrafoAsset.replace("/", "-")
            strTrafoAsset = strTrafoAsset.replace("\\", "-")
            trafo_status_folder = target_NeRDA_folder + "\\" + "Transformer_Status"
            if not os.path.isdir(trafo_status_folder):
                os.mkdir(trafo_status_folder)
            dfInService_trafo.to_excel(trafo_status_folder + "\\" + strTrafoAsset + ".xlsx")
        print("Transformers completed...")
        pb=pb+1
        pbar.updateProgressBar(pb+1, "Debug Mode: Outputting SyncMachine Loading Data Files")
        for strSynmAsset, dfInService_synm in dictInService_synm.items():
            strSynmAsset = strSynmAsset.replace("/", "-")
            strSynmAsset = strSynmAsset.replace("\\", "-")
            synm_status_folder = target_NeRDA_folder + "\\" + "Synmachine_Status"
            if not os.path.isdir(synm_status_folder):
                os.mkdir(synm_status_folder)
            dfInService_synm.to_excel(synm_status_folder + "\\" + strSynmAsset + ".xlsx")
        print("Synchronous machines completed...")
        pb=pb+1
        pbar.updateProgressBar(pb+1, "Debug Mode: Outputting Load Loading Data Files")
        for strLoadAsset, dfInService_load in dictInService_load.items():
            strLoadAsset = strLoadAsset.replace("/", "-")
            strLoadAsset = strLoadAsset.replace("\\", "-")
            load_status_folder = target_NeRDA_folder + "\\" + "Load_Status"
            if not os.path.isdir(load_status_folder):
                os.mkdir(load_status_folder)
            dfInService_load.to_excel(load_status_folder + "\\" + strLoadAsset + ".xlsx")
        print(f"Lines, transformers and synchronous machine' in-service status data succesfully saved in {target_NeRDA_folder}...")
        print("################################################################")
    return status, msg, dictLineLoading, dictTrafoLoading, dfDownScalingFactor_grid, dfAllTrafos_grid_PQnom, \
           dictLinePowerBus1, dictLineIBus1, dictLineIBus2, dictTrafoPowerLv, dictTrafoIBus1, dictTrafoIBus2

def PF_CalcOverLoading(pf, strHHmode, type, dictLoadingAssets, threshold, start_date, PSA_running_mode, dfAsset,
                       dir_line_loading_data, dir_trafo_loading_data, dictLinePowerBus1, dictLineIBus1, dictLineIBus2,
                       dictTrafoPowerLv, dictTrafoIBus1, dictTrafoIBus2, number_of_days, number_of_half_hours):
    
    Time = pd.Series(
        ['%s:%s' % (h, m) for h in ([00] + list(range(1, int(number_of_half_hours / 2)))) for m in ('00', '30')])
    Days = pd.Series(['%s%s' % ('Day', d) for d in range(number_of_days)])

    status_msg = ""
    combinebEvent = 0
    lCombineEvents = []
    lCombineConstraints = []
    strPSA_running_mode = ""
    threshold_new = threshold
    dictAssetRating = dict()
    dictAssetBSPPrimaryFeeder = defaultdict(lambda: defaultdict(dict))

    if type == "trafo":
        trafos = pf.app.GetCalcRelevantObjects('*.ElmTr2')
        for index, row in dfAsset.iterrows():
            if const.strAssetTrans in row['ASSET_TYPE']:
                for trafo in trafos:
                    trafoname = trafo.loc_name
                    if trafoname == row['ASSET_ID']:
                        dictAssetRating[trafo.loc_name] = row['MAX_LOADING']
                        dictAssetBSPPrimaryFeeder[trafo.loc_name]['BSP'] = row['BSP']
                        dictAssetBSPPrimaryFeeder[trafo.loc_name]['PRIMARY'] = row['PRIMARY']
                        dictAssetBSPPrimaryFeeder[trafo.loc_name]['FEEDER'] = row['FEEDER']
    if type == "line":
        lines = pf.app.GetCalcRelevantObjects('*.ElmLne')
        for index, row in dfAsset.iterrows():
            if const.strAssetLine in row['ASSET_TYPE']:
                for line in lines:
                    linename = line.loc_name
                    if linename == row['ASSET_ID']:
                        dictAssetRating[line.loc_name] = row['MAX_LOADING']
                        dictAssetBSPPrimaryFeeder[line.loc_name]['BSP'] = row['BSP']
                        dictAssetBSPPrimaryFeeder[line.loc_name]['PRIMARY'] = row['PRIMARY']
                        dictAssetBSPPrimaryFeeder[line.loc_name]['FEEDER'] = row['FEEDER']
    for strNameAsset, dfLoadingAsset in dictLoadingAssets.items():
        if strNameAsset in dictAssetRating.keys():
            threshold_new = dictAssetRating[strNameAsset]
        bEvent, msg, df_event = validation.checkOverload(strHHmode=strHHmode, df=dfLoadingAsset,
                                                         threshold=threshold_new, start_date=start_date)
        combinebEvent += bEvent
        if pfconf.bOutputAllLoading:
            bOutputOverLoading = False
            # dfLoadingAsset.set_axis(Days, axis=1, inplace=True)
            # dfLoadingAsset.set_axis(Time, axis=0, inplace=True)
            dfLoadingAsset = dfLoadingAsset.set_axis(Days, axis=1)
            dfLoadingAsset = dfLoadingAsset.set_axis(Time, axis=0)
            if type == "line":
                if "/" in strNameAsset:
                    strNameAsset = strNameAsset.replace("/", "-")
                if not os.path.isdir(dir_line_loading_data):
                    os.mkdir(dir_line_loading_data)
                dfLoadingAsset.to_excel(
                    dir_line_loading_data + "PSA_" + PSA_running_mode + "_Loading-" + strNameAsset + ".xlsx")
            if type == "trafo":
                if "/" in strNameAsset:
                    strNameAsset = strNameAsset.replace("/", "-")
                if not os.path.isdir(dir_trafo_loading_data):
                    os.mkdir(dir_trafo_loading_data)
                dfLoadingAsset.to_excel(
                    dir_trafo_loading_data + "PSA_" + PSA_running_mode + "_Loading-" + strNameAsset + ".xlsx")
        else:
            bOutputOverLoading = True
        if bEvent:
            if bOutputOverLoading:
                # dfLoadingAsset.set_axis(Days,axis=1,inplace=True)
                # dfLoadingAsset.set_axis(Time, axis=0,inplace=True)
                dfLoadingAsset = dfLoadingAsset.set_axis(Days, axis=1)
                dfLoadingAsset = dfLoadingAsset.set_axis(Time, axis=0)
                if type == "line":
                    if "/" in strNameAsset:
                        strNameAsset = strNameAsset.replace("/", "-")
                    if not os.path.isdir(dir_line_loading_data):
                        os.mkdir(dir_line_loading_data)
                    # PF_createXl(dir_line_loading_data, "PSA_" + PSA_running_mode + "_Loading-" + strNameAsset, "DATA")
                    dfLoadingAsset.to_excel(
                        dir_line_loading_data + "PSA_" + PSA_running_mode + "_Loading-" + strNameAsset + ".xlsx")
                if type == "trafo":
                    if "/" in strNameAsset:
                        strNameAsset = strNameAsset.replace("/", "-")
                    if not os.path.isdir(dir_trafo_loading_data):
                        os.mkdir(dir_trafo_loading_data)
                    # PF_createXl(dir_trafo_loading_data, "PSA_" + PSA_running_mode + "_Loading-" + strNameAsset, "DATA")
                    dfLoadingAsset.to_excel(
                        dir_trafo_loading_data + "PSA_" + PSA_running_mode + "_Loading-" + strNameAsset + ".xlsx")
            lAssetID = [strNameAsset for i in range(df_event.shape[0])]
            if PSA_running_mode == "B":
                strPSA_running_mode = "BASE"
            elif PSA_running_mode == "M":
                strPSA_running_mode = "MAINT"
            elif PSA_running_mode == "C":
                strPSA_running_mode = "CONT"
            elif PSA_running_mode == "MC":
                strPSA_running_mode = "MAINT_CONT"
            lScenario = [strPSA_running_mode for i in range(df_event.shape[0])]
            if type == "line":
                oAsset = pf.app.GetCalcRelevantObjects(strNameAsset + '.ElmLne')[0]
                lSecondary = ["NONE" for i in range(df_event.shape[0])]
                # lRating = [0 if oAsset.Inom_a == None or oAsset.Unom == None else oAsset.Inom_a * oAsset.Unom * 1000 for i in
                #            range(df_event.shape[0])]
            elif type == "trafo":
                oAsset = pf.app.GetCalcRelevantObjects(strNameAsset + '.ElmTr2')[0]
                lSecondary = [strNameAsset if oAsset.typ_id.utrn_h < 30 and oAsset.typ_id.utrn_h > 10 else "NONE" for i
                              in range(df_event.shape[0])]
                # lRating = [0 if oAsset.Snom_a == None else oAsset.Snom_a for i in range(df_event.shape[0])]
                # lRating = [0 if oAsset.bushv.cterm.uknom == None or oAsset.Inom_h_a == None else oAsset.bushv.cterm.
                #                                              uknom * oAsset.Inom_h_a for i in range(df_event.shape[0])]
            else:
                lSecondary = ["NONE" for i in range(df_event.shape[0])]

            # lRating = [oAsset.pRating for i in range(df_event.shape[0])]
            # lType = [type for i in range(df_event.shape[0])]
            # lRating = [0 if oAsset.pRating==None else oAsset.pRating for i in range(df_event.shape[0])]
            # lRating = [0 if oAsset.maxload == None else oAsset.maxload for i in range(df_event.shape[0])]
            lBSP = ["COLO" for i in range(df_event.shape[0])]
            lPrimary = [dictAssetBSPPrimaryFeeder[strNameAsset]['PRIMARY'] for i in range(df_event.shape[0])]
            lFeeder = [dictAssetBSPPrimaryFeeder[strNameAsset]['FEEDER'] for i in range(df_event.shape[0])]
            # lPrimary = [oAsset.cpGrid.loc_name.split('-')[0].split('_')[0] for i in range(df_event.shape[0])]
            # lFeeder = [oAsset.cpGrid.loc_name for i in range(df_event.shape[0])]
            lAssetType = [oAsset.GetClassName() for i in range(df_event.shape[0])]
            # lThreshold = [threshold_new for i in range(df_event.shape[0])]
            lPower = []
            lI_nom = []
            lI_nom_adj = []
            # lLoading = []

            if type == "line":
                if pfconf.bDebug:
                    print("**************** strNameAsset = " + strNameAsset)
                ElmLne = pf.app.GetCalcRelevantObjects(strNameAsset + '.ElmLne')[0]
                # lne_from_bus = ElmLne.bus1.cterm.loc_name
                # lne_to_bus = ElmLne.bus2.cterm.loc_name
                # lLoading = ElmLne.GetAttribute('c:loading')
                # lFromBus = [lne_from_bus for i in range(df_event.shape[0])]
                # lToBus = [lne_to_bus for i in range(df_event.shape[0])]

                for start_time in df_event["start_time"].tolist():
                    col, row = findcolrow(str(start_time), str(start_date))
                    sign = getSign(dictLinePowerBus1[strNameAsset].iloc[row, col])  # the sign will be positive if the bus 1 power is positive
                    I_bus1 = dictLineIBus1[strNameAsset].iloc[row, col] # JPO: Is this equivalent to  line.GetAttribute('m:I:bus1')? - yes
                    I_bus2 = dictLineIBus2[strNameAsset].iloc[row, col] # JPO: Is this equivalent to line.GetAttribute('m:I:bus2')? - yes
                    if I_bus1 > I_bus2:
                        I_bus = I_bus1
                        if ElmLne.outserv == 0 and ElmLne.HasResults() == 1:
                            U_term = ElmLne.Unom * ElmLne.GetAttribute('m:u:bus1')  # kV * p.u.
                        else:
                            U_term = 0
                    else:
                        I_bus = I_bus2
                        if ElmLne.outserv == 0 and ElmLne.HasResults() == 1:
                            U_term = ElmLne.Unom * ElmLne.GetAttribute('m:u:bus2')
                        else:
                            U_term = 0
                    I_nom=ElmLne.Inom_a
                    I_nom_adj = I_nom * threshold_new / 100
                    I_exceed_kA = I_bus - I_nom_adj
                    if I_exceed_kA < 0:
                        I_exceed_kA = 0
                    power = sign * np.sqrt(3) * I_exceed_kA * U_term * 1000  # positive if SPM, negative if SEPM
                    lPower.append(power)
                    lI_nom.append(I_nom)
                    lI_nom_adj.append(I_nom_adj)


            if type == "trafo":
                ElmTr2 = pf.app.GetCalcRelevantObjects(strNameAsset + '.ElmTr2')[0]
                tr2_from_bus = ElmTr2.bushv.cterm.loc_name
                tr2_to_bus = ElmTr2.buslv.cterm.loc_name
                # tr2_loading = ElmTr2.GetAttribute('c:loading')
                lToBus = [tr2_to_bus for i in range(df_event.shape[0])]
                lFromBus = [tr2_from_bus for i in range(df_event.shape[0])]
                # lLoading = [tr2_loading for i in range(df_event.shape[0])]
                for start_time in df_event["start_time"].tolist():
                    col, row = findcolrow(str(start_time), str(start_date))
                    sign = getSign(dictTrafoPowerLv[strNameAsset].iloc[row, col])  # the sign is positive if lv bus power is positive
                    I_bus1 = dictTrafoIBus1[strNameAsset].iloc[row, col]  # lv
                    I_bus2 = dictTrafoIBus2[strNameAsset].iloc[row, col]  # hv

                    if sign > 0:  # SEPM
                        I_bus = I_bus1
                        if ElmTr2.outserv == 0 and ElmTr2.HasResults() == 1:
                            U_term = ElmTr2.buslv.cterm.uknom * ElmTr2.buslv.cterm.GetAttribute('m:u')  # kV * p.u.
                            i_nom = ElmTr2.GetAttribute('e:Inom_l_a')
                        else:
                            U_term = 0
                            i_nom = 0
                    else:  # SPM
                        I_bus = I_bus2
                        if ElmTr2.outserv == 0 and ElmTr2.HasResults() == 1:
                            U_term = ElmTr2.bushv.cterm.uknom * ElmTr2.bushv.cterm.GetAttribute('m:u')
                            i_nom = ElmTr2.GetAttribute('e:Inom_h_a')
                        else:
                            U_term = 0
                            i_nom = 0


                    # if I_bus1 > I_bus2:
                    #     I_bus = I_bus1
                    #     if ElmTr2.outserv == 0 and ElmTr2.HasResults() == 1:
                    #         U_term = ElmTr2.buslv.cterm.uknom * ElmTr2.buslv.cterm.GetAttribute('m:u')  # kV * p.u.
                    #         i_nom = ElmTr2.GetAttribute('e:Inom_l_a')
                    #     else:
                    #         U_term = 0
                    #         i_nom = 0
                    # else:
                    #     I_bus = I_bus2
                    #     if ElmTr2.outserv == 0 and ElmTr2.HasResults() == 1:
                    #         U_term = ElmTr2.bushv.cterm.uknom * ElmTr2.bushv.cterm.GetAttribute('m:u')
                    #         i_nom = ElmTr2.GetAttribute('e:Inom_h_a')
                    #     else:
                    #         U_term = 0
                    #         i_nom = 0

                    I_nom_adj = i_nom * threshold_new / 100
                    I_exceed_kA = I_bus - I_nom_adj
                    power = - sign * I_exceed_kA * U_term * np.sqrt(3) * 1e3
                    lPower.append(power)
                    lI_nom.append(i_nom)
                    lI_nom_adj.append(I_nom_adj)

            else:
                lFromBus = ["None" for i in range(df_event.shape[0])]
                lToBus = ["None" for i in range(df_event.shape[0])]

            # lAssetType = "999"

            lReqID = [i for i in range(df_event.shape[0])]
            df_constraints = df_event
            # df_event = df_event.drop('maximum_loading', axis=1)
            df_event = df_event.rename(columns={'maximum_loading': 'loading_pct'})
            df_event.insert(0, 'req_id', lReqID)
            df_event.insert(1, 'bsp', lBSP)
            df_event.insert(2, 'primary', lPrimary)
            df_event.insert(3, 'feeder', lFeeder)
            df_event.insert(4, 'secondary', lSecondary)
            df_event.insert(5, 'terminal', lToBus)
            df_event.insert(6, 'constrained_pf_id', lAssetID)
            df_event.insert(7, 'constrained_pf_type', lAssetType)
            df_event.insert(8, 'required_power_kw', lPower)
            df_event.insert(9, 'busbar_from', lFromBus)
            df_event.insert(10, 'busbar_to', lToBus)
            df_event.insert(11, 'scenario', lScenario)
            # df_event.insert(14, 'loading_pct', lLoading)
            lCombineEvents.append(df_event)

            df_constraints = df_constraints.rename(columns={'maximum_loading': 'loading_pct'})
            df_constraints.insert(0, 'req_id', lReqID)
            df_constraints.insert(1, 'bsp', lBSP)
            df_constraints.insert(2, 'primary', lPrimary)
            df_constraints.insert(3, 'feeder', lFeeder)
            df_constraints.insert(4, 'secondary', lSecondary)
            df_constraints.insert(5, 'terminal', lToBus)
            df_constraints.insert(6, 'constrained_pf_id', lAssetID)
            df_constraints.insert(7, 'constrained_pf_type', lAssetType)
            df_constraints.insert(12, 'nominal_current_kA', lI_nom)
            df_constraints.insert(13, 'nominal_current_adj_kA', lI_nom_adj)
            df_constraints.insert(14, 'scenario', lScenario)
            lCombineConstraints.append(df_constraints)

            threshold_new = threshold

    if combinebEvent > 0:
        bEvents = True
        if type == "line":
            status_msg = "Overloaded events recorded for Lines"
        if type == "trafo":
            status_msg = "Overloaded events recorded for Transformers"

    else:
        bEvents = False
        if type == "line":
            status_msg = "No overloaded events found for Lines at given threshold"
        if type == "trafo":
            status_msg = "No overloaded events found for Transformers at given threshold"

    return status_msg, lCombineEvents, lCombineConstraints, bEvents


def PF_CombineEvents(bLineEvents, lCombinedLineEvents, bTrafoEvents, lCombinedTrafoEvents):
    if bLineEvents and not bTrafoEvents:
        lCombinedEvents = lCombinedLineEvents
        dfCombinedEvents = pd.concat(lCombinedEvents)
    elif bTrafoEvents and not bLineEvents:
        lCombinedEvents = lCombinedTrafoEvents
        dfCombinedEvents = pd.concat(lCombinedEvents)
    elif bLineEvents and bTrafoEvents:
        lCombinedEvents = lCombinedLineEvents + lCombinedTrafoEvents
        dfCombinedEvents = pd.concat(lCombinedEvents)
    else:
        # dfCombinedEvents = pd.DataFrame(data=["No overloading events for lines and transformers."])
        dfCombinedEvents = pd.DataFrame()
    return dfCombinedEvents

def PF_CombineConstraints(bLineEvents, lCombinedLineConstraints, bTrafoEvents, lCombinedTrafoConstraints):
    if bLineEvents and not bTrafoEvents:
        lCombinedEvents = lCombinedLineConstraints
        dfCombinedConstraints = pd.concat(lCombinedEvents)
    elif bTrafoEvents and not bLineEvents:
        lCombinedEvents = lCombinedTrafoConstraints
        dfCombinedConstraints = pd.concat(lCombinedEvents)
    elif bLineEvents and bTrafoEvents:
        lCombinedEvents = lCombinedLineConstraints + lCombinedTrafoConstraints
        dfCombinedConstraints = pd.concat(lCombinedEvents)
    else:
        # dfCombinedEvents = pd.DataFrame(data=["No overloading events for lines and transformers."])
        dfCombinedConstraints = pd.DataFrame()
    return dfCombinedConstraints

##############
### OUTPUT ###
##############
def PF_OutputConstraints(bLineEvents, bTrafoEvents, dfCombinedEvents, dir_output, PSArunID):
    if not os.path.isdir(dir_output):
        os.mkdir(dir_output)
    if not bLineEvents and not bTrafoEvents:
        dfCombinedEvents.reset_index(drop=True, inplace=True)
        dfCombinedEvents.to_excel(dir_output + "\\" + PSArunID + const.strPSANoConstraints + ".xlsx", index=False)
    else:
        dfCombinedEvents.reset_index(drop=True, inplace=True)
        dfCombinedEvents.to_excel(dir_output + "\\" + PSArunID + const.strPSAConstraints + ".xlsx", index=False)

    return

def PF_OutputFlexReqts(bLineEvents, bTrafoEvents, dfCombinedEvents, dir_output, PSArunID):
    if not os.path.isdir(dir_output):
        os.mkdir(dir_output)
    if not bLineEvents and not bTrafoEvents:
        dfCombinedEvents.reset_index(drop=True, inplace=True)
        dfCombinedEvents.to_excel(dir_output + "\\" + PSArunID + const.strPSANoFlexReqts + ".xlsx", index=False)
    else:
        dfCombinedEvents.reset_index(drop=True, inplace=True)
        dfCombinedEvents.to_excel(dir_output + "\\" + PSArunID + const.strPSAFlexReqts + ".xlsx", index=False)

    return


# def PF_CalcOverLoading(type, dir_input,dir_output,threshold,start_date):
#     status_msg = ""
#     combinebEvent = 0
#     if not os.path.isdir(dir_input):
#         os.mkdir(dir_input)
#     if not os.path.isdir(dir_output):
#         os.mkdir(dir_output)
#     for file in os.listdir(dir_input):
#         if file.endswith(".xlsx"):
#             bEvent, msg, df_event = validation.checkOverload(file_path=dir_input+"\\"+file,threshold=threshold,start_date=start_date)
#             combinebEvent += bEvent
#             if bEvent:
#                 df_event.to_excel(dir_output+"\\"+file)
#             else:
#                 df_event.to_excel(dir_output+"\\"+file, index=False, header=False)
#
#     if combinebEvent > 0:
#         if type == "line":
#             status_msg = "Overloaded events recorded for lines in {}".format(dir_output)
#         if type == "trafo":
#             status_msg = "Overloaded events recorded for trafos in {}".format(dir_output)
#
#     else:
#         if type == "line":
#             status_msg = "No overloaded events found for lines at given threshold {}".format(threshold)
#         if type == "trafo":
#             status_msg = "No overloaded events found for trafos at given threshold {}".format(threshold)
#
#     return status_msg

# pf.prepare_loadflow(pfconf.ldf_mode, pfconf.algorithm, pfconf.trafo_tap, pfconf.shunt_tap, pfconf.zip_load,
#                     pfconf.q_limits, pfconf.phase_shift, pfconf.trafo_tap_limit, pfconf.max_iter)
#
# PF_updateConnectivity(pf,no_days=10,no_half_hours=48,lOutageAssets=lOutageAssets, lOutageAssetTypes=lOutageAssetTypes,
#                       lDfOutages=lDfOutages)

# dir_tgt_lne = "C:\\Wentao Zhu\\Project\\PSA\\Test Data\\Test Input Files\\Line Loading Table"
# dir_evt_lne = "C:\\Wentao Zhu\\Project\\PSA\\Test Data\\Test Input Files\\Line Events"
# for file in os.listdir(dir_tgt_lne):
#     if file.endswith(".xlsx"):
#         msg, df_event = validation.checkOverload(file_path=dir_tgt_lne+"\\"+file,threshold=50,start_date=start_date)
#         df_event.to_excel(dir_evt_lne+"\\"+file)
#
# dir_tgt_trafo = "C:\\Wentao Zhu\\Project\\PSA\\Test Data\\Test Input Files\\Trafo Loading Table"
# dir_evt_trafo = "C:\\Wentao Zhu\\Project\\PSA\\Test Data\\Test Input Files\\Trafo Events"
# for file in os.listdir(dir_tgt_trafo):
#     if file.endswith(".xlsx"):
#         msg, df_event = validation.checkOverload(file_path=dir_tgt_trafo+"\\"+file,threshold=50,start_date=start_date)
#         df_event.to_excel(dir_evt_trafo+"\\"+file)

def PF_createGen(pf, name_gen, gen_terminal):
    pf.createGen(name_gen=name_gen, gen_terminal=gen_terminal)
    return

def PF_outAssetLoading(target_folder):
    """
    read asset (line/tx) %loading from individual half-hour period file and paste into an array
    """
    pass